import uuid
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from sqlalchemy import and_
from database.models import Order
from database.session import SessionLocal


def generate_range_report(start_date, end_date):
    db = SessionLocal()
    file_name = f"report_{uuid.uuid4()}.xlsx"

    try:
        orders = db.query(Order).filter(
            and_(
                Order.created_at >= start_date,
                Order.created_at <= end_date,
            )
        ).order_by(Order.created_at).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        bold = Font(bold=True)

        current_day = None
        daily_total = 0
        total = 0

        for order in orders:
            order_day = order.created_at.date()

            if current_day != order_day:
                if current_day is not None:
                    ws.append(["", "", "Kun jami:", daily_total])
                    ws.append([])

                ws.append([f"Sana: {order_day.strftime('%Y-%m-%d')}"])
                ws.append(["ID", "Mashina", "Xizmat", "Narx", "Moykachi", "Vaqt"])
                for col in range(1, 7):
                    ws.cell(row=ws.max_row, column=col).font = bold

                current_day = order_day
                daily_total = 0

            price = order.price or 0
            daily_total += price
            total += price

            ws.append([
                order.id,
                order.car.plate_number if order.car else "",
                order.services_name or "",
                price,
                order.washer.full_name if order.washer else "",
                order.created_at.strftime("%H:%M")
            ])

        if current_day is not None:
            ws.append(["", "", "Kun jami:", daily_total])
            ws.append([])

        ws.append(["", "", "JAMI:", total])
        ws.cell(row=ws.max_row, column=3).font = bold
        ws.cell(row=ws.max_row, column=4).font = bold

        wb.save(file_name)
        return file_name

    finally:
        db.close()


def generate_monthly_report(year: int, month: int):
    db = SessionLocal()
    file_name = f"monthly_{uuid.uuid4()}.xlsx"

    try:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)

        orders = db.query(Order).filter(
            and_(
                Order.created_at >= start_date,
                Order.created_at < end_date,
            )
        ).order_by(Order.created_at).all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"{month}-{year}"
        bold = Font(bold=True)

        current_day = None
        daily_total = 0
        monthly_total = 0

        for order in orders:
            order_day = order.created_at.date()

            if current_day != order_day:
                if current_day is not None:
                    ws.append(["", "", "Kun jami:", daily_total])
                    ws.append([])

                ws.append([f"Sana: {order_day.strftime('%Y-%m-%d')}"])
                ws.append(["ID", "Mashina", "Xizmat", "Narx", "Moykachi", "Vaqt"])
                for col in range(1, 7):
                    ws.cell(row=ws.max_row, column=col).font = bold

                current_day = order_day
                daily_total = 0

            price = order.price or 0
            daily_total += price
            monthly_total += price

            ws.append([
                order.id,
                order.car.plate_number if order.car else "",
                order.services_name or "",
                price,
                order.washer.full_name if order.washer else "",
                order.created_at.strftime("%H:%M")
            ])

        if current_day is not None:
            ws.append(["", "", "Kun jami:", daily_total])
            ws.append([])

        ws.append(["", "", "OYLIK JAMI:", monthly_total])
        ws.cell(row=ws.max_row, column=3).font = bold
        ws.cell(row=ws.max_row, column=4).font = bold

        wb.save(file_name)
        return file_name

    finally:
        db.close()